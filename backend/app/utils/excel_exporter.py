from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from typing import List, Dict, Any, Optional
from io import BytesIO
from sqlalchemy.orm import Session
from app.repositories.inventory import (
    StockInRepository,
    StockOutRepository,
    InstalledRepository,
    ReturnRepository,
    MaterialRepository,
    MandorRepository,
)
from app.utils.formatters import format_date_indonesia


# Styles
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(size=10)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def create_excel_export(
    db: Session,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    mandor_id: Optional[int] = None,
    material_id: Optional[int] = None,
    search: Optional[str] = None,
) -> BytesIO:
    """
    Create Excel file dengan multiple sheets untuk semua data inventory
    Returns BytesIO object
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Initialize repositories
    stock_in_repo = StockInRepository(db)
    stock_out_repo = StockOutRepository(db)
    installed_repo = InstalledRepository(db)
    return_repo = ReturnRepository(db)
    material_repo = MaterialRepository(db)
    mandor_repo = MandorRepository(db)
    
    # Set default date range jika tidak ada
    if not start_date:
        start_date = date.today().replace(day=1)  # Start of month
    if not end_date:
        end_date = date.today()
    
    # Sheet 1: Barang Masuk
    create_stock_in_sheet(
        wb, stock_in_repo, material_repo, start_date, end_date, material_id
    )
    
    # Sheet 2: Barang Keluar
    create_stock_out_sheet(
        wb, stock_out_repo, material_repo, mandor_repo, start_date, end_date, mandor_id, material_id
    )
    
    # Sheet 3: Barang Terpasang
    create_installed_sheet(
        wb, installed_repo, material_repo, mandor_repo, start_date, end_date, mandor_id, material_id
    )
    
    # Sheet 4: Barang Pengembalian
    create_return_sheet(
        wb, return_repo, material_repo, mandor_repo, start_date, end_date, mandor_id, material_id
    )
    
    # Sheet 5: Summary
    create_summary_sheet(
        wb, stock_in_repo, stock_out_repo, installed_repo, return_repo,
        material_repo, start_date, end_date, search
    )
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_stock_in_sheet(
    wb: Workbook,
    repo: StockInRepository,
    material_repo: MaterialRepository,
    start_date: date,
    end_date: date,
    material_id: Optional[int] = None
):
    """Create sheet for Barang Masuk"""
    ws = wb.create_sheet("Barang Masuk")
    
    headers = [
        "No", "Tanggal Masuk", "Nomor Invoice", "Kode Barang", 
        "Nama Barang", "Satuan", "Quantity", "Created At"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
    
    # Get data
    stock_ins = repo.get_by_date_range(start_date, end_date, skip=0, limit=10000)
    if material_id:
        stock_ins = [si for si in stock_ins if si.material_id == material_id]
    
    # Write data
    for idx, stock_in in enumerate(stock_ins, 2):
        material = material_repo.get(stock_in.material_id)
        ws.cell(row=idx, column=1, value=idx - 1).border = BORDER
        ws.cell(row=idx, column=2, value=format_date_indonesia(stock_in.tanggal_masuk)).border = BORDER
        ws.cell(row=idx, column=3, value=stock_in.nomor_invoice).border = BORDER
        ws.cell(row=idx, column=4, value=material.kode_barang if material else "").border = BORDER
        ws.cell(row=idx, column=5, value=material.nama_barang if material else "").border = BORDER
        ws.cell(row=idx, column=6, value=material.satuan if material else "").border = BORDER
        ws.cell(row=idx, column=7, value=stock_in.quantity).border = BORDER
        ws.cell(row=idx, column=8, value=format_date_indonesia(stock_in.created_at.date())).border = BORDER
    
    # Auto width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def create_stock_out_sheet(
    wb: Workbook,
    repo: StockOutRepository,
    material_repo: MaterialRepository,
    mandor_repo: MandorRepository,
    start_date: date,
    end_date: date,
    mandor_id: Optional[int] = None,
    material_id: Optional[int] = None
):
    """Create sheet for Barang Keluar"""
    ws = wb.create_sheet("Barang Keluar")
    
    headers = [
        "No", "Tanggal Keluar", "Nomor Barang Keluar", "Mandor",
        "Kode Barang", "Nama Barang", "Satuan", "Quantity", "Created At"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
    
    # Get data
    stock_outs = repo.get_by_date_range(start_date, end_date, skip=0, limit=10000)
    if mandor_id:
        stock_outs = [so for so in stock_outs if so.mandor_id == mandor_id]
    if material_id:
        stock_outs = [so for so in stock_outs if so.material_id == material_id]
    
    # Write data
    for idx, stock_out in enumerate(stock_outs, 2):
        material = material_repo.get(stock_out.material_id)
        mandor = mandor_repo.get(stock_out.mandor_id)
        ws.cell(row=idx, column=1, value=idx - 1).border = BORDER
        ws.cell(row=idx, column=2, value=format_date_indonesia(stock_out.tanggal_keluar)).border = BORDER
        ws.cell(row=idx, column=3, value=stock_out.nomor_barang_keluar).border = BORDER
        ws.cell(row=idx, column=4, value=mandor.nama if mandor else "").border = BORDER
        ws.cell(row=idx, column=5, value=material.kode_barang if material else "").border = BORDER
        ws.cell(row=idx, column=6, value=material.nama_barang if material else "").border = BORDER
        ws.cell(row=idx, column=7, value=material.satuan if material else "").border = BORDER
        ws.cell(row=idx, column=8, value=stock_out.quantity).border = BORDER
        ws.cell(row=idx, column=9, value=format_date_indonesia(stock_out.created_at.date())).border = BORDER
    
    # Auto width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def create_installed_sheet(
    wb: Workbook,
    repo: InstalledRepository,
    material_repo: MaterialRepository,
    mandor_repo: MandorRepository,
    start_date: date,
    end_date: date,
    mandor_id: Optional[int] = None,
    material_id: Optional[int] = None
):
    """Create sheet for Barang Terpasang"""
    ws = wb.create_sheet("Barang Terpasang")
    
    headers = [
        "No", "Tanggal Pasang", "Mandor", "Kode Barang",
        "Nama Barang", "Satuan", "Quantity", "Created At"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
    
    # Get data
    installed_items = repo.get_by_date_range(start_date, end_date, skip=0, limit=10000)
    if mandor_id:
        installed_items = [i for i in installed_items if i.mandor_id == mandor_id]
    if material_id:
        installed_items = [i for i in installed_items if i.material_id == material_id]
    
    # Write data
    for idx, installed in enumerate(installed_items, 2):
        material = material_repo.get(installed.material_id)
        mandor = mandor_repo.get(installed.mandor_id)
        ws.cell(row=idx, column=1, value=idx - 1).border = BORDER
        ws.cell(row=idx, column=2, value=format_date_indonesia(installed.tanggal_pasang)).border = BORDER
        ws.cell(row=idx, column=3, value=mandor.nama if mandor else "").border = BORDER
        ws.cell(row=idx, column=4, value=material.kode_barang if material else "").border = BORDER
        ws.cell(row=idx, column=5, value=material.nama_barang if material else "").border = BORDER
        ws.cell(row=idx, column=6, value=material.satuan if material else "").border = BORDER
        ws.cell(row=idx, column=7, value=installed.quantity).border = BORDER
        ws.cell(row=idx, column=8, value=format_date_indonesia(installed.created_at.date())).border = BORDER
    
    # Auto width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def create_return_sheet(
    wb: Workbook,
    repo: ReturnRepository,
    material_repo: MaterialRepository,
    mandor_repo: MandorRepository,
    start_date: date,
    end_date: date,
    mandor_id: Optional[int] = None,
    material_id: Optional[int] = None
):
    """Create sheet for Barang Pengembalian"""
    ws = wb.create_sheet("Barang Pengembalian")
    
    headers = [
        "No", "Tanggal Kembali", "Mandor", "Kode Barang",
        "Nama Barang", "Satuan", "Quantity Kembali", "Quantity Kondisi Baik", "Quantity Kondisi Reject", "Created At"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
    
    # Get data
    returns = repo.get_by_date_range(start_date, end_date, skip=0, limit=10000)
    if mandor_id:
        returns = [r for r in returns if r.mandor_id == mandor_id]
    if material_id:
        returns = [r for r in returns if r.material_id == material_id]
    
    # Write data
    for idx, return_item in enumerate(returns, 2):
        material = material_repo.get(return_item.material_id)
        mandor = mandor_repo.get(return_item.mandor_id)
        ws.cell(row=idx, column=1, value=idx - 1).border = BORDER
        ws.cell(row=idx, column=2, value=format_date_indonesia(return_item.tanggal_kembali)).border = BORDER
        ws.cell(row=idx, column=3, value=mandor.nama if mandor else "").border = BORDER
        ws.cell(row=idx, column=4, value=material.kode_barang if material else "").border = BORDER
        ws.cell(row=idx, column=5, value=material.nama_barang if material else "").border = BORDER
        ws.cell(row=idx, column=6, value=material.satuan if material else "").border = BORDER
        ws.cell(row=idx, column=7, value=return_item.quantity_kembali).border = BORDER
        ws.cell(row=idx, column=8, value=return_item.quantity_kondisi_baik or 0).border = BORDER
        ws.cell(row=idx, column=9, value=return_item.quantity_kondisi_reject or 0).border = BORDER
        ws.cell(row=idx, column=10, value=format_date_indonesia(return_item.created_at.date())).border = BORDER
    
    # Auto width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def create_summary_sheet(
    wb: Workbook,
    stock_in_repo: StockInRepository,
    stock_out_repo: StockOutRepository,
    installed_repo: InstalledRepository,
    return_repo: ReturnRepository,
    material_repo: MaterialRepository,
    start_date: date,
    end_date: date,
    search: Optional[str] = None
):
    """Create summary sheet"""
    ws = wb.create_sheet("Summary")
    
    headers = [
        "Kode Barang", "Nama Barang", "Satuan", 
        "Barang Masuk", "Barang Keluar", "Barang Terpasang", 
        "Barang Kembali", "Total Kondisi Baik", "Total Kondisi Reject", "Stok Ready", "Stock Saat Ini"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
    
    # Get all materials
    materials = material_repo.get_active_materials(skip=0, limit=10000)
    if search:
        s = (search or '').strip().lower()
        materials = [m for m in materials if (m.kode_barang or '').lower().find(s) != -1 or (m.nama_barang or '').lower().find(s) != -1]
    
    # Write data
    for idx, material in enumerate(materials, 2):
        # Calculate totals
        stock_ins = stock_in_repo.get_by_material(material.id, skip=0, limit=10000)
        stock_outs = [so for so in stock_out_repo.get_by_date_range(start_date, end_date, 0, 10000) 
                     if so.material_id == material.id]
        installed = [i for i in installed_repo.get_by_date_range(start_date, end_date, 0, 10000)
                    if i.material_id == material.id]
        returns = [r for r in return_repo.get_by_date_range(start_date, end_date, 0, 10000)
                  if r.material_id == material.id]
        
        total_in = sum(si.quantity for si in stock_ins)
        total_out = sum(so.quantity for so in stock_outs)
        total_installed = sum(i.quantity for i in installed)
        total_return = sum(r.quantity_kembali for r in returns)
        # Get returns yang sudah dikeluarkan kembali (is_released = 1) untuk retur keluar
        returns_released = [r for r in returns if getattr(r, 'is_released', 0) == 1]
        total_retur_keluar = sum(r.quantity_kembali for r in returns_released)
        total_kondisi_baik = sum((r.quantity_kondisi_baik or 0) for r in returns)
        total_kondisi_reject = sum((r.quantity_kondisi_reject or 0) for r in returns)
        
        # Rumus Stock Balance yang Dinamis & Best Practice (sync dengan stock_service.py)
        # Keluar Efektif = Keluar - Retur Keluar (retur keluar mengurangi efek keluar karena sudah kembali ke gudang)
        keluar_efektif = total_out - total_retur_keluar
        # Stock Saat Ini = Masuk - Keluar Efektif + Kembali
        current_stock = total_in - keluar_efektif + total_return
        # Stok Ready = Stock Saat Ini - Kondisi Reject (barang reject tidak bisa dikeluarkan)
        stock_ready = max(0, current_stock - total_kondisi_reject)
        
        ws.cell(row=idx, column=1, value=material.kode_barang).border = BORDER
        ws.cell(row=idx, column=2, value=material.nama_barang).border = BORDER
        ws.cell(row=idx, column=3, value=material.satuan).border = BORDER
        ws.cell(row=idx, column=4, value=total_in).border = BORDER
        ws.cell(row=idx, column=5, value=total_out).border = BORDER
        ws.cell(row=idx, column=6, value=total_installed).border = BORDER
        ws.cell(row=idx, column=7, value=total_return).border = BORDER
        ws.cell(row=idx, column=8, value=total_kondisi_baik).border = BORDER
        ws.cell(row=idx, column=9, value=total_kondisi_reject).border = BORDER
        ws.cell(row=idx, column=10, value=stock_ready).border = BORDER
        ws.cell(row=idx, column=11, value=current_stock).border = BORDER
    
    # Auto width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

