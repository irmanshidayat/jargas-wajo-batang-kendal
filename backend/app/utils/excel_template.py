from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


# Styles untuk template
HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")  # Blue-500
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(size=10)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def create_material_import_template() -> BytesIO:
    """
    Create template Excel file untuk import materials
    Format: NO, NAMA BARANG, KODE BARANG, SATUAN, KATEGORI, HARGA
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Import Material"
    
    # Headers
    headers = ["NO", "NAMA BARANG", "KODE BARANG", "SATUAN", "KATEGORI", "HARGA"]
    
    # Write headers di row 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
    
    # Contoh data (row 2-4)
    example_data = [
        [1, "PIPA PE 125mm", "PE125", "M", "PIPA DITRIBUSI", 150000],
        [2, "PIPA PE 63MM", "PE63", "M", "PIPA DITRIBUSI", 120000],
        [3, "EQUALTEE 63mm", "EQ63", "PCS", "PIPA DITRIBUSI", 85000],
        [4, "Regulator RT", "REG-RT", "PCS", "BUNGAN RUMAH", 250000],
        [5, "Meter Rumah Tangga G 1.6", "MTR-RT", "PCS", "BUNGAN RUMAH", 350000],
        [6, "Ball Valves 3/4\" Gas Specification", "BV-34", "PCS", "BUNGAN KOMPOR", 180000],
        [7, "Slang 1/2\"", "SLG-12", "M", "BUNGAN KOMPOR", 45000],
    ]
    
    # Write example data
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.border = BORDER
            if col_idx == 1:  # NO column
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN
    
    # Set column widths
    column_widths = {
        'A': 8,   # NO
        'B': 40,  # NAMA BARANG
        'C': 20,  # KODE BARANG
        'D': 12,  # SATUAN
        'E': 20,  # KATEGORI
        'F': 15,  # HARGA
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Add note/info di bawah data (optional)
    info_row = len(example_data) + 3
    ws.merge_cells(f'A{info_row}:F{info_row + 3}')
    info_cell = ws.cell(row=info_row, column=1)
    info_cell.value = (
        "PETUNJUK:\n"
        "1. Jangan hapus atau ubah header di row pertama\n"
        "2. Isi data mulai dari row kedua\n"
        "3. Kolom KODE BARANG, KATEGORI, dan HARGA boleh dikosongkan\n"
        "4. Kategori yang valid: PIPA DITRIBUSI, BUNGAN RUMAH, BUNGAN KOMPOR\n"
        "5. HARGA harus berupa angka (contoh: 150000 atau 150000.50)"
    )
    info_cell.font = Font(size=9, italic=True, color="666666")
    info_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

