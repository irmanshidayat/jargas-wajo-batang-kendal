from openpyxl import load_workbook
from typing import List, Dict, Optional, Tuple
from io import BytesIO
from app.core.exceptions import ValidationError


def parse_excel_file(
    file_content: bytes,
    expected_headers: List[str],
    start_row: int = 2,
    skip_empty_rows: bool = True
) -> Tuple[List[Dict[str, any]], List[str]]:
    """
    Parse Excel file dan return list of dictionaries dan list of errors
    
    Args:
        file_content: Bytes content dari uploaded file
        expected_headers: List of expected column headers di row pertama
        start_row: Row number dimana data mulai (1-indexed, default 2 = skip header)
        skip_empty_rows: Jika True, skip rows yang semua kolomnya kosong
    
    Returns:
        Tuple of (list of data dictionaries, list of error messages)
    """
    errors = []
    data = []
    
    try:
        # Load workbook dari bytes
        workbook = load_workbook(filename=BytesIO(file_content), read_only=False, data_only=True)
        
        # Ambil sheet pertama
        if not workbook.sheetnames:
            raise ValidationError("File Excel tidak memiliki sheet")
        
        worksheet = workbook[workbook.sheetnames[0]]
        
        # Validasi header (row 1)
        if worksheet.max_row < 1:
            raise ValidationError("File Excel kosong")
        
        header_row = []
        for col in range(1, len(expected_headers) + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            header_row.append(str(cell_value).strip() if cell_value else "")
        
        # Validasi header sesuai dengan expected headers (case insensitive)
        header_map = {}
        for idx, expected in enumerate(expected_headers):
            found = False
            for i, actual in enumerate(header_row):
                if actual and expected.lower() in actual.lower() or actual.lower() in expected.lower():
                    header_map[expected] = i + 1  # Excel column (1-indexed)
                    found = True
                    break
            if not found:
                raise ValidationError(f"Header '{expected}' tidak ditemukan di file Excel")
        
        # Parse data rows
        for row_num in range(start_row, worksheet.max_row + 1):
            row_data = {}
            is_empty = True
            
            # Extract data per kolom
            for header, col_num in header_map.items():
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    # Convert to string dan strip whitespace
                    cell_value = str(cell_value).strip()
                    if cell_value:
                        is_empty = False
                else:
                    cell_value = ""
                
                # Map header ke key (lowercase dengan underscore)
                key = header.lower().replace(" ", "_")
                row_data[key] = cell_value
            
            # Skip empty rows jika diminta
            if skip_empty_rows and is_empty:
                continue
            
            # Validasi minimal data tidak kosong (harus ada isi)
            if not is_empty:
                # Tambahkan row number untuk error reporting
                row_data['_row_number'] = row_num
                data.append(row_data)
        
        workbook.close()
        
    except ValidationError:
        raise
    except Exception as e:
        raise ValidationError(f"Gagal membaca file Excel: {str(e)}")
    
    return data, errors


def validate_material_row(
    row: Dict[str, any],
    valid_kategoris: Optional[List[str]] = None
) -> Tuple[bool, Optional[str]]:
    """
    Validate single material row data
    
    Args:
        row: Dictionary dengan keys: no, nama_barang, kode_barang, satuan, kategori
        valid_kategoris: List of valid kategori values (case insensitive)
    
    Returns:
        Tuple of (is_valid, error_message)
    """
    row_num = row.get('_row_number', '?')
    
    # Validasi nama_barang (wajib)
    nama_barang = row.get('nama_barang', '').strip()
    if not nama_barang:
        return False, f"Row {row_num}: Nama Barang wajib diisi"
    
    if len(nama_barang) > 255:
        return False, f"Row {row_num}: Nama Barang maksimal 255 karakter"
    
    # Validasi satuan (wajib)
    satuan = row.get('satuan', '').strip()
    if not satuan:
        return False, f"Row {row_num}: Satuan wajib diisi"
    
    if len(satuan) > 50:
        return False, f"Row {row_num}: Satuan maksimal 50 karakter"
    
    # Validasi kode_barang (opsional, tapi kalau ada harus valid)
    kode_barang = row.get('kode_barang', '').strip()
    if kode_barang and len(kode_barang) > 100:
        return False, f"Row {row_num}: Kode Barang maksimal 100 karakter"
    
    # Validasi kategori (opsional, tapi kalau ada harus dari list valid)
    kategori = row.get('kategori', '').strip()
    if kategori:
        if len(kategori) > 100:
            return False, f"Row {row_num}: Kategori maksimal 100 karakter"
        
        if valid_kategoris:
            # Case insensitive comparison
            kategori_upper = kategori.upper().strip()
            valid_upper = [k.upper() for k in valid_kategoris]
            if kategori_upper not in valid_upper:
                return False, f"Row {row_num}: Kategori '{kategori}' tidak valid. Kategori yang valid: {', '.join(valid_kategoris)}"
    
    return True, None

